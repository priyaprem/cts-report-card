import openpyxl
import mysql.connector

path = "Rank_Table.xlsx"
wb = openpyxl.load_workbook(path)
total_college_month_event_id = []
mark_table = []

cts = mysql.connector.connect(
  host="localhost",
  user="root",
  password="root",
  database="cts"
)

cts_cursor = cts.cursor()

for ws in wb:
    max_row = ws.max_row + 1 
    max_col = ws.max_column + 1
    college_month_event_id = []
    for i in range (2, max_col):
        college_month_event_id.append(ws.cell(1,i).value)
        total_college_month_event_id.append(ws.cell(1,i).value)
    marks = []
    for i in range(2, max_row):
        student_id = ws.cell(i,1).value
        for j in range(2, max_col):
            mark = [student_id, college_month_event_id[j-2], ws.cell(i,j).value]
            marks.append(mark)
            insert_sql = "insert into mark(student_id, college_month_event_id, mark) values (%s, %s, %s)"
            cts_cursor.execute(insert_sql, tuple(mark))
    
cts.commit()      

    
